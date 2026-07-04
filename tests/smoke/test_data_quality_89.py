from pathlib import Path

import openpyxl

from main import parse_args
from parser.excel import export
from parser.ranking_guard import apply_ranking_guard
from parser.ranking_power_sanity_guard import PowerRecoveryCandidate, _pending_placeholder


def _ranking_row(rank: int, power: int, name: str, alliance: str = "SWSq") -> dict:
    return {
        "rank": rank,
        "ocr_rank": rank,
        "computed_rank": rank,
        "player_name": name,
        "name": name,
        "alliance_tag": alliance,
        "power": power,
        "hero_power": power,
        "source_file": "server553_thp_window_10_15.png",
    }


def test_data_quality_validation_defaults_to_no_cache_even_when_compat_flag_is_absent():
    args = parse_args(["--mode", "development"])
    assert args.mode == "development"
    assert args.ocr_cache is False
    assert args.no_ocr_cache is False


def test_ranking_guard_quarantine_preserves_visible_rank_slot_and_following_ranks():
    grouped = {
        (553, "total_hero_power"): [
            {"rank": 10, "ocr_rank": 10, "computed_rank": 10, "name": "", "player_name": "", "alliance_tag": "SWSq", "power": 203_127_000, "source_file": "server553_thp_window_10_15.png"},
            {"rank": 11, "ocr_rank": 11, "computed_rank": 11, "name": "Next Wolf", "player_name": "Next Wolf", "alliance_tag": "SWSq", "power": 202_991_100, "source_file": "server553_thp_window_10_15.png"},
            {"rank": 12, "ocr_rank": 12, "computed_rank": 12, "name": "Anchor Wolf", "player_name": "Anchor Wolf", "alliance_tag": "SWSq", "power": 202_500_100, "source_file": "server553_thp_window_10_15.png"},
        ]
    }

    guarded = apply_ranking_guard(grouped)
    rows = guarded[(553, "total_hero_power")]
    pending = rows[0]

    assert pending["rank"] == 10
    assert pending["ocr_rank"] == 10
    assert pending["pending_review"] is True
    assert pending["rank_slot_preserved"] is True
    assert rows[1]["rank"] == 11
    assert rows[2]["rank"] == 12
    assert [row["rank"] for row in rows] == [10, 11, 12]
    assert ("REVIEW", "ranking_guard_quarantine") in guarded


def test_power_placeholder_preserves_sven_raw_display_for_review():
    row = {
        "rank": 10,
        "ocr_rank": 10,
        "computed_rank": 10,
        "name": "Sven the vän",
        "player_name": "Sven the vän",
        "alliance_tag": "SWSq",
        "power": 20_312_700,
        "source_file": "server553_thp_window_10_15.png",
    }
    candidate = PowerRecoveryCandidate(203_127_000, 0.91, ["scale_x10_truncated_digit"], 0.93)

    pending = _pending_placeholder(row, ranking_type="total_hero_power", reason="ambiguous", candidates=[candidate])

    assert pending["rank"] == 10
    assert pending["observed_name"] == "Sven the vän"
    assert pending["observed_alliance"] == "SWSq"
    assert pending["rank_slot_preserved"] is True
    assert pending["power_sort_anchor"] == 203_127_000


def test_export_surfaces_pending_slot_and_raw_display_fidelity(tmp_path: Path):
    out = tmp_path / "rank_slot_export.xlsx"
    grouped = {
        (553, "alliance_power"): [
            {
                "rank": 10,
                "ocr_rank": 10,
                "computed_rank": 10,
                "pending_review": True,
                "pending_review_reason": "thp_power_outlier",
                "rank_slot_preserved": True,
                "observed_name": "Sven the vän",
                "normalized_name": "sven the van",
                "canonical_name": "",
                "observed_alliance": "SWSq",
                "normalized_alliance": "swsq",
                "canonical_alliance": "",
                "alliance_tag": "SWSq",
                "player_name": "Sven the vän",
                "name": "PENDING REVIEW | Sven the vän",
                "power": 203_127_000,
                "source_file": "server553_thp_window_10_15.png",
            },
            {
                "rank": 11,
                "ocr_rank": 11,
                "computed_rank": 11,
                "alliance_tag": "SWSq",
                "player_name": "Next Wolf",
                "name": "Next Wolf",
                "power": 202_991_100,
                "source_file": "server553_thp_window_10_15.png",
            },
        ]
    }

    export(grouped, filename=out)
    wb = openpyxl.load_workbook(out)
    ws = wb["553_alliance_power"]
    headers = [cell.value for cell in ws[1]]
    row_10 = {headers[index]: value for index, value in enumerate(next(ws.iter_rows(min_row=2, max_row=2, values_only=True)))}
    row_11 = {headers[index]: value for index, value in enumerate(next(ws.iter_rows(min_row=3, max_row=3, values_only=True)))}

    assert row_10["rank"] == 10
    assert row_10["pending_review"] is True
    assert row_10["rank_slot_preserved"] is True
    assert row_10["observed_name"] == "Sven the vän"
    assert row_10["observed_alliance"] == "SWSq"
    assert row_11["rank"] == 11
